"""
Excel Document Agent — handles the full lifecycle for .xlsx / .xls files.

Uses openpyxl for reading and writing Excel workbooks. Maps worksheets
to CanvasNode trees (heading + table per sheet) and reconstructs
workbooks from CanvasBundle data on export.
"""

from __future__ import annotations

import re
import uuid
from io import BytesIO
from typing import Any, Optional

from .base import (
    AgentCapability,
    AtomGroup,
    CanvasBundle,
    CanvasNode,
    ComplianceConstraints,
    ExportResult,
    DocumentAgent,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Sheet-name patterns used during atomization to infer volume categories.
_CATEGORY_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"budget|cost|price|pricing", re.IGNORECASE), "cost_volume"),
    (re.compile(r"personnel|staff|team|labor", re.IGNORECASE), "key_personnel"),
    (re.compile(r"schedule|timeline|gantt|milestones?", re.IGNORECASE), "schedule"),
    (re.compile(r"risk", re.IGNORECASE), "risk_register"),
    (re.compile(r"requirement|compliance|matrix|trace", re.IGNORECASE), "compliance_matrix"),
    (re.compile(r"equipment|material|asset", re.IGNORECASE), "equipment_list"),
    (re.compile(r"summary|overview|executive", re.IGNORECASE), "executive_summary"),
]

_MAX_INGEST_ROWS = 100
MAX_INGEST_SIZE = 100 * 1024 * 1024  # 100 MB


def _require_openpyxl() -> None:
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel document processing. "
            "Install it with: pip install openpyxl"
        )


def _cell_style_dict(cell: Any) -> dict[str, Any]:
    """Extract formatting info from an openpyxl cell into a plain dict."""
    style: dict[str, Any] = {}
    font = cell.font
    if font:
        if font.bold:
            style["bold"] = True
        if font.italic:
            style["italic"] = True
        if font.underline and font.underline != "none":
            style["underline"] = True
        if font.size:
            style["font_size"] = font.size
        if font.name:
            style["font_name"] = font.name
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            style["font_color"] = str(font.color.rgb)
    alignment = cell.alignment
    if alignment:
        if alignment.horizontal:
            style["align"] = alignment.horizontal
        if alignment.vertical:
            style["valign"] = alignment.vertical
        if alignment.wrap_text:
            style["wrap_text"] = True
    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format
    return style


def _infer_category(sheet_name: str) -> str:
    """Infer a proposal category from a worksheet name."""
    for pattern, category in _CATEGORY_PATTERNS:
        if pattern.search(sheet_name):
            return category
    return "general"


def _tags_from_sheet(sheet_name: str) -> list[str]:
    """Generate suggested library tags from a sheet name."""
    tags = ["spreadsheet"]
    cat = _infer_category(sheet_name)
    if cat != "general":
        tags.append(cat)
    # Add the lowercase sheet name as a tag.
    normalised = re.sub(r"[^a-z0-9]+", "_", sheet_name.lower()).strip("_")
    if normalised and normalised not in tags:
        tags.append(normalised)
    return tags


class XlsxAgent(DocumentAgent):
    """Document agent for Microsoft Excel (.xlsx / .xls) files."""

    format_id = "xlsx"
    display_name = "Excel Document Agent"
    file_extensions = ["xlsx", "xls"]
    capabilities = {
        AgentCapability.READ_NATIVE,
        AgentCapability.WRITE_NATIVE,
        AgentCapability.FORMULAS,
        AgentCapability.CONDITIONAL_FORMATTING,
        AgentCapability.NAMED_RANGES,
        AgentCapability.CHARTS,
        AgentCapability.PDF_RENDER,
    }

    def __init__(self) -> None:
        _require_openpyxl()

    # ── Ingest ────────────────────────────────────────────────────────

    async def ingest(self, file_bytes: bytes, filename: str) -> CanvasBundle:
        """Read an Excel file and produce a CanvasBundle."""
        if not file_bytes:
            raise ValueError(f"XlsxAgent: empty file '{filename}'")
        if len(file_bytes) > MAX_INGEST_SIZE:
            raise ValueError(
                f"XlsxAgent: file '{filename}' exceeds {MAX_INGEST_SIZE // (1024 * 1024)}MB limit"
            )

        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True)
        except Exception as exc:
            raise ValueError(f"XlsxAgent: failed to parse '{filename}' — {exc}") from exc

        try:
            nodes: list[CanvasNode] = []

            for sheet_idx, ws in enumerate(wb.worksheets):
                # ── Heading node for the sheet ──
                nodes.append(
                    CanvasNode(
                        id=str(uuid.uuid4()),
                        type="heading",
                        content={"text": ws.title, "level": 1},
                        provenance={"source": filename, "sheet": ws.title},
                    )
                )

                # ── Build merged-cell lookup ──
                # Maps (min_row, min_col) -> {"rowSpan": n, "colSpan": n}
                merged_map: dict[tuple[int, int], dict[str, int]] = {}
                # Set of cells that are *inside* a merge but not the top-left anchor
                merged_interior: set[tuple[int, int]] = set()
                for merge_range in ws.merged_cells.ranges:
                    min_row = merge_range.min_row
                    min_col = merge_range.min_col
                    max_row = merge_range.max_row
                    max_col = merge_range.max_col
                    row_span = max_row - min_row + 1
                    col_span = max_col - min_col + 1
                    merged_map[(min_row, min_col)] = {
                        "rowSpan": row_span,
                        "colSpan": col_span,
                    }
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            if (r, c) != (min_row, min_col):
                                merged_interior.add((r, c))

                # ── Read rows ──
                all_rows: list[list[dict[str, Any]]] = []
                row_count = 0
                truncated = False

                for row in ws.iter_rows():
                    row_count += 1
                    if row_count > _MAX_INGEST_ROWS:
                        truncated = True
                        break
                    cells: list[dict[str, Any]] = []
                    for cell in row:
                        r, c = cell.row, cell.column
                        # Skip interior merged cells — only the anchor carries data
                        if (r, c) in merged_interior:
                            continue
                        value = cell.value
                        cell_dict: dict[str, Any] = {
                            "value": value if value is not None else "",
                        }
                        # Attach merge spans
                        span = merged_map.get((r, c))
                        if span:
                            if span["colSpan"] > 1:
                                cell_dict["colSpan"] = span["colSpan"]
                            if span["rowSpan"] > 1:
                                cell_dict["rowSpan"] = span["rowSpan"]
                        # Attach cell style
                        style = _cell_style_dict(cell)
                        if style:
                            cell_dict["style"] = style
                        cells.append(cell_dict)
                    all_rows.append(cells)

                # ── Separate headers from data ──
                headers: list[dict[str, Any]] = []
                data_rows: list[list[dict[str, Any]]] = []
                if all_rows:
                    headers = all_rows[0]
                    data_rows = all_rows[1:]

                # ── Column widths ──
                column_widths: list[Optional[float]] = []
                if ws.column_dimensions:
                    max_col_idx = ws.max_column or 0
                    for col_idx in range(1, max_col_idx + 1):
                        col_letter = get_column_letter(col_idx)
                        dim = ws.column_dimensions.get(col_letter)
                        if dim and dim.width:
                            column_widths.append(dim.width)
                        else:
                            column_widths.append(None)

                # ── Table node ──
                table_meta: dict[str, Any] = {}
                if truncated:
                    actual_max = ws.max_row or row_count
                    table_meta["truncated"] = True
                    table_meta["total_rows"] = actual_max
                    table_meta["note"] = (
                        f"Sheet contains {actual_max} rows; only the first "
                        f"{_MAX_INGEST_ROWS} are included."
                    )

                nodes.append(
                    CanvasNode(
                        id=str(uuid.uuid4()),
                        type="table",
                        content={
                            "headers": headers,
                            "rows": data_rows,
                            "column_widths": column_widths,
                        },
                        style={},
                        provenance={
                            "source": filename,
                            "sheet": ws.title,
                        },
                        library_tags=_tags_from_sheet(ws.title),
                    )
                )

                if table_meta:
                    nodes[-1].content["metadata"] = table_meta

                # ── Page break between sheets ──
                if sheet_idx < len(wb.worksheets) - 1:
                    nodes.append(
                        CanvasNode(
                            id=str(uuid.uuid4()),
                            type="page_break",
                            content={},
                        )
                    )

            # ── Workbook-level metadata ──
            meta: dict[str, Any] = {"filename": filename}
            props = wb.properties
            if props:
                if props.title:
                    meta["title"] = props.title
                if props.creator:
                    meta["creator"] = props.creator
                if props.subject:
                    meta["subject"] = props.subject
                if props.keywords:
                    meta["keywords"] = props.keywords
                if props.created:
                    meta["created"] = str(props.created)
                if props.modified:
                    meta["modified"] = str(props.modified)

            return CanvasBundle(
                document_id=str(uuid.uuid4()),
                nodes=nodes,
                constraints=ComplianceConstraints(format_type="letter"),
                metadata=meta,
                source_format="xlsx",
                source_agent=self.display_name,
            )
        finally:
            wb.close()

    # ── Atomize ───────────────────────────────────────────────────────

    async def atomize(self, bundle: CanvasBundle) -> list[AtomGroup]:
        """Break a CanvasBundle into per-worksheet AtomGroups."""
        groups: list[AtomGroup] = []
        current_nodes: list[CanvasNode] = []
        current_heading: Optional[str] = None

        for node in bundle.nodes:
            if node.type == "page_break":
                # Flush the accumulated sheet group
                if current_nodes:
                    category = _infer_category(current_heading or "")
                    groups.append(
                        AtomGroup(
                            nodes=list(current_nodes),
                            heading_text=current_heading,
                            suggested_category=category,
                            suggested_tags=_tags_from_sheet(current_heading or "Sheet"),
                            confidence=0.8 if category != "general" else 0.5,
                            source_sheet=current_heading,
                        )
                    )
                current_nodes = []
                current_heading = None
                continue

            if node.type == "heading" and node.content.get("level") == 1:
                current_heading = node.content.get("text", "")

            current_nodes.append(node)

        # Flush the last group (after the last sheet, there is no trailing page_break)
        if current_nodes:
            category = _infer_category(current_heading or "")
            groups.append(
                AtomGroup(
                    nodes=list(current_nodes),
                    heading_text=current_heading,
                    suggested_category=category,
                    suggested_tags=_tags_from_sheet(current_heading or "Sheet"),
                    confidence=0.8 if category != "general" else 0.5,
                    source_sheet=current_heading,
                )
            )

        return groups

    # ── Export ─────────────────────────────────────────────────────────

    async def export(self, bundle: CanvasBundle) -> ExportResult:
        """Render a CanvasBundle back to an Excel workbook."""
        _require_openpyxl()

        wb = Workbook()
        # Remove the default sheet created by openpyxl
        if wb.worksheets:
            wb.remove(wb.worksheets[0])

        warnings: list[str] = []
        constraints = bundle.constraints
        current_ws = None
        table_col_count = 1  # track width for text_block merges

        for node in bundle.nodes:
            # ── New worksheet from level-1 heading ──
            if node.type == "heading" and node.content.get("level") == 1:
                sheet_name = str(node.content.get("text", "Sheet"))[:31]
                current_ws = wb.create_sheet(title=sheet_name)
                table_col_count = 1
                continue

            # Ensure we have a worksheet to write into
            if current_ws is None:
                current_ws = wb.create_sheet(title="Sheet1")
                table_col_count = 1

            # ── Table node ──
            if node.type == "table":
                headers = node.content.get("headers", [])
                rows = node.content.get("rows", [])
                column_widths = node.content.get("column_widths", [])

                # Determine default font from constraints
                default_font_name = constraints.font_family or "Calibri"
                default_font_size = constraints.font_size or 11

                # Write headers (row 1)
                for col_idx, header_cell in enumerate(headers, start=1):
                    cell = current_ws.cell(row=1, column=col_idx)
                    cell.value = header_cell.get("value", "")
                    # Header style: bold by default
                    header_style = header_cell.get("style", {})
                    cell.font = Font(
                        name=header_style.get("font_name", default_font_name),
                        size=header_style.get("font_size", default_font_size),
                        bold=header_style.get("bold", True),
                        italic=header_style.get("italic", False),
                    )
                    cell.alignment = Alignment(
                        horizontal=header_style.get("align", "left"),
                        vertical=header_style.get("valign", "bottom"),
                        wrap_text=header_style.get("wrap_text", False),
                    )
                    # Apply header border (thin bottom)
                    cell.border = Border(
                        bottom=Side(style="thin"),
                    )
                    # Handle merged header cells
                    col_span = header_cell.get("colSpan", 1)
                    row_span = header_cell.get("rowSpan", 1)
                    if col_span > 1 or row_span > 1:
                        current_ws.merge_cells(
                            start_row=1,
                            start_column=col_idx,
                            end_row=1 + row_span - 1,
                            end_column=col_idx + col_span - 1,
                        )

                table_col_count = max(table_col_count, len(headers))

                # Write data rows (starting at row 2)
                for row_offset, row_data in enumerate(rows, start=2):
                    for col_idx, cell_data in enumerate(row_data, start=1):
                        cell = current_ws.cell(row=row_offset, column=col_idx)
                        cell.value = cell_data.get("value", "")

                        cell_style = cell_data.get("style", {})
                        cell.font = Font(
                            name=cell_style.get("font_name", default_font_name),
                            size=cell_style.get("font_size", default_font_size),
                            bold=cell_style.get("bold", False),
                            italic=cell_style.get("italic", False),
                        )
                        cell.alignment = Alignment(
                            horizontal=cell_style.get("align", "left"),
                            vertical=cell_style.get("valign", "bottom"),
                            wrap_text=cell_style.get("wrap_text", False),
                        )

                        # Number format
                        nf = cell_style.get("number_format")
                        if nf:
                            cell.number_format = nf

                        # Borders — light grid
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

                        # Merged cells
                        col_span = cell_data.get("colSpan", 1)
                        row_span = cell_data.get("rowSpan", 1)
                        if col_span > 1 or row_span > 1:
                            current_ws.merge_cells(
                                start_row=row_offset,
                                start_column=col_idx,
                                end_row=row_offset + row_span - 1,
                                end_column=col_idx + col_span - 1,
                            )

                    table_col_count = max(table_col_count, len(row_data))

                # Apply column widths
                for cw_idx, width in enumerate(column_widths):
                    if width is not None:
                        col_letter = get_column_letter(cw_idx + 1)
                        current_ws.column_dimensions[col_letter].width = width

            # ── Text block → merged cell spanning table width ──
            elif node.type == "text_block":
                text = node.content.get("text", "")
                next_row = (current_ws.max_row or 0) + 1
                cell = current_ws.cell(row=next_row, column=1)
                cell.value = text
                cell.alignment = Alignment(wrap_text=True)
                span = max(table_col_count, 1)
                if span > 1:
                    current_ws.merge_cells(
                        start_row=next_row,
                        start_column=1,
                        end_row=next_row,
                        end_column=span,
                    )

            # ── Page break — no action needed (sheets already separated) ──
            elif node.type == "page_break":
                continue

        # If the workbook ended up empty (no nodes), add at least one sheet
        if not wb.worksheets:
            wb.create_sheet(title="Sheet1")

        # ── Print area and page setup ──
        for ws in wb.worksheets:
            if ws.max_row and ws.max_column:
                last_col = get_column_letter(ws.max_column)
                ws.print_area = f"A1:{last_col}{ws.max_row}"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            if constraints.margins:
                ws.page_margins.top = constraints.margins.get("top", 0.75) / 72
                ws.page_margins.bottom = constraints.margins.get("bottom", 0.75) / 72
                ws.page_margins.left = constraints.margins.get("left", 0.7) / 72
                ws.page_margins.right = constraints.margins.get("right", 0.7) / 72

        # ── Apply font defaults from constraints across all cells ──
        if constraints.font_family or constraints.font_size:
            target_name = constraints.font_family or "Calibri"
            target_size = constraints.font_size or 11
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.font:
                            cell.font = Font(
                                name=target_name,
                                size=target_size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                underline=cell.font.underline,
                                color=cell.font.color,
                            )

        # ── Serialize ──
        sheet_count = len(wb.sheetnames)
        buf = BytesIO()
        wb.save(buf)
        wb.close()

        original_name = bundle.metadata.get("filename", "export.xlsx")
        base_name = original_name.rsplit(".", 1)[0] if "." in original_name else original_name
        export_filename = f"{base_name}.xlsx"

        return ExportResult(
            file_bytes=buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=export_filename,
            format="xlsx",
            page_count=sheet_count,
            warnings=warnings,
        )
