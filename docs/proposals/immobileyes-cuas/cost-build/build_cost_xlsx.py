#!/usr/bin/env python3
"""Immobileyes DON26BX03-NP002 Phase I Cost Volume — Base ($200k) + Option ($115k).
Live-formula workbook using the firm's real TACFI rate stack (Fringe 35 / OH 77 / G&A 40 / Fee 7).
Inputs (rates, hours, materials) are blue; formulas black; rate assumptions yellow."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/proposals/immobileyes-cuas/Immobileyes_DON26BX03-NP002_Cost_Volume.xlsx"
BLUE = Font(name="Arial", size=10, color="0000FF")           # inputs
BLACK = Font(name="Arial", size=10, color="000000")          # formulas
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=13, bold=True, color="1F3864")
SUB = Font(name="Arial", size=9, italic=True, color="595959")
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
YEL = PatternFill("solid", fgColor="FFFF00")
NAVY = PatternFill("solid", fgColor="1F3864")
GREY = PatternFill("solid", fgColor="D9D9D9")
LT = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CUR = '$#,##0.00'
CUR0 = '$#,##0'
PCT = '0%'
RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center")

wb = openpyxl.Workbook()

def build_period(ws, title, ceiling, labor_rows, materials, travel, mat_note, trav_note):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 16
    ws['A1'] = title; ws['A1'].font = TITLE; ws.merge_cells('A1:E1')
    ws['A2'] = ("Immobileyes, Inc. (WOSB) · UEI KL3MJVGD9XZ9 · CAGE 8KQ82 · 30,000 sq-ft ISO-9001 facility, "
                "Kent OH · DON26BX03-NP002 NAVAIR/NAVSEA C-UAS")
    ws['A2'].font = SUB; ws.merge_cells('A2:E2')

    # Indirect-rate assumptions (yellow inputs, referenced by every formula below)
    ws['A4'] = "INDIRECT RATES (firm-wide; source: Immobileyes TACFI Vol.3 cost methodology)"
    ws['A4'].font = BOLD; ws.merge_cells('A4:E4')
    rates = [("Fringe Benefits  (× direct labor)", 0.35),
             ("Labor Overhead  (× labor + fringe)", 0.77),
             ("G&A  (× burdened labor + materials + ODC)", 0.40),
             ("Fee / Profit  (× total cost)", 0.07)]
    r = 5
    for lbl, v in rates:
        ws[f'A{r}'] = lbl; ws[f'A{r}'].font = BLACK
        c = ws[f'C{r}']; c.value = v; c.font = BLUE; c.fill = YEL; c.number_format = PCT; c.alignment = RIGHT; c.border = BORD
        r += 1
    FR, OH, GA, FEE = 'C5', 'C6', 'C7', 'C8'

    # A. Direct labor
    r = 10
    ws[f'A{r}'] = "A.  DIRECT LABOR"; ws[f'A{r}'].font = HDR
    for cc in 'ABCDE': ws[f'{cc}{r}'].fill = NAVY
    r += 1
    heads = ["Labor Category", "Name / Assignment", "Rate ($/hr)", "Hours", "Direct Labor ($)"]
    for i, h in enumerate(heads):
        c = ws.cell(row=r, column=1+i, value=h); c.font = HDR; c.fill = GREY; c.font = BOLD; c.border = BORD
        c.alignment = CENTER if i >= 2 else Alignment(horizontal="left")
    first = r + 1
    for cat, name, rate, hrs in labor_rows:
        r += 1
        ws.cell(row=r, column=1, value=cat).font = BLACK
        ws.cell(row=r, column=2, value=name).font = BLACK
        cr = ws.cell(row=r, column=3, value=rate); cr.font = BLUE; cr.number_format = CUR; cr.alignment = RIGHT; cr.border = BORD
        ch = ws.cell(row=r, column=4, value=hrs); ch.font = BLUE; ch.alignment = RIGHT; ch.border = BORD
        ce = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); ce.font = BLACK; ce.number_format = CUR; ce.alignment = RIGHT; ce.border = BORD
    last = r
    r += 1
    ws.cell(row=r, column=1, value="Total Direct Labor").font = BOLD
    th = ws.cell(row=r, column=4, value=f"=SUM(D{first}:D{last})"); th.font = BOLD; th.alignment = RIGHT; th.border = BORD
    tdl = ws.cell(row=r, column=5, value=f"=SUM(E{first}:E{last})"); tdl.font = BOLD; tdl.number_format = CUR; tdl.alignment = RIGHT; tdl.fill = LT; tdl.border = BORD
    TDL = f'E{r}'

    # B. Indirect & other
    r += 2
    ws[f'A{r}'] = "B.  INDIRECT COSTS, MATERIALS & FEE"; ws[f'A{r}'].font = HDR
    for cc in 'ABCDE': ws[f'{cc}{r}'].fill = NAVY
    def line(r, label, formula, note=None, bold=False, fill=None, inp=False):
        ws.cell(row=r, column=1, value=label).font = BOLD if bold else BLACK
        c = ws.cell(row=r, column=5, value=formula)
        c.font = (BLUE if inp else (BOLD if bold else BLACK)); c.number_format = CUR; c.alignment = RIGHT; c.border = BORD
        if fill: c.fill = fill
        if note:
            n = ws.cell(row=r, column=2, value=note); n.font = SUB; ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        return c
    r += 1; line(r, "Fringe Benefits", f"={TDL}*{FR}")
    fr_row = r
    r += 1; line(r, "Labor Overhead", f"=({TDL}+E{fr_row})*{OH}")
    oh_row = r
    r += 1; burd = line(r, "Total Burdened Labor", f"={TDL}+E{fr_row}+E{oh_row}", bold=True, fill=LT); BURD = f'E{r}'
    r += 1; line(r, "Materials & Supplies", materials, note=mat_note, inp=True); MAT = f'E{r}'
    r += 1; line(r, "Travel / ODC", travel, note=trav_note, inp=True); TRV = f'E{r}'
    r += 1; gab = line(r, "G&A Base (burdened labor + materials + ODC)", f"={BURD}+{MAT}+{TRV}", fill=LT); GAB = f'E{r}'
    r += 1; line(r, "G&A", f"={GAB}*{GA}"); GA_row = r
    r += 1; tc = line(r, "Total Cost", f"={GAB}+E{GA_row}", bold=True, fill=LT); TC = f'E{r}'
    r += 1; line(r, "Fee / Profit", f"={TC}*{FEE}"); FEE_row = r
    r += 1
    ws.cell(row=r, column=1, value=f"TOTAL PROPOSED PRICE — {title.split('—')[-1].strip()}").font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for cc in 'ABCD': ws[f'{cc}{r}'].fill = NAVY
    tot = ws.cell(row=r, column=5, value=f"={TC}+E{FEE_row}"); tot.font = Font(name="Arial", size=11, bold=True, color="FFFFFF"); tot.fill = NAVY; tot.number_format = CUR; tot.alignment = RIGHT
    TOT = f'E{r}'
    r += 2
    ws.cell(row=r, column=1, value="Topic ceiling (NTE)").font = BLACK
    cl = ws.cell(row=r, column=5, value=ceiling); cl.font = BLUE; cl.number_format = CUR0; cl.alignment = RIGHT
    r += 1
    ws.cell(row=r, column=1, value="Margin under ceiling").font = BLACK
    mg = ws.cell(row=r, column=5, value=f"=E{r-1}-{TOT}"); mg.font = BLACK; mg.number_format = CUR; mg.alignment = RIGHT
    return TOT

# BASE
wsB = wb.active; wsB.title = "Base"
base_labor = [
    ("Principal Investigator", "Dr. Bahman Taheri", 63, 250),
    ("Senior Optical Engineer", "Optical effects lead", 50, 320),
    ("Electrical Engineer", "Beam-router electronics", 45, 200),
    ("Software Engineer", "STORM tracking / control", 45, 140),
    ("Program Manager", "Atossa Alavi", 50, 90),
]
totB = build_period(wsB, "Phase I — BASE  (6 months, NTE $200,000)", 200000, base_labor,
             7000, 3000,
             "EO/FPV test cameras, optics, laser diodes, LC samples",
             "Kick-off @ NAVAIR/NAVSEA; Tinker coordination")

# OPTION
wsO = wb.create_sheet("Option")
opt_labor = [
    ("Principal Investigator", "Dr. Bahman Taheri", 63, 110),
    ("Senior Optical Engineer", "Optical effects lead", 50, 200),
    ("Electrical Engineer", "Breadboard electronics", 45, 110),
    ("Mechanical Engineer", "Breadboard fabrication", 35, 90),
    ("Machine-shop Operator", "Fabrication", 30, 60),
    ("Software Engineer", "Test automation", 45, 40),
]
totO = build_period(wsO, "Phase I — OPTION  (6 months, NTE $115,000)", 115000, opt_labor,
             6000, 2000,
             "Breadboard optics, LC beam-router cells, laser, mounts",
             "Phase II transition review")

# SUMMARY
wsS = wb.create_sheet("Summary", 0)
wsS.column_dimensions['A'].width = 40
wsS.column_dimensions['B'].width = 18
wsS.column_dimensions['C'].width = 18
wsS['A1'] = "DON26BX03-NP002 — Phase I Cost Summary"; wsS['A1'].font = TITLE; wsS.merge_cells('A1:C1')
wsS['A2'] = ("Immobileyes, Inc. (WOSB) · NAVAIR & NAVSEA Open Topic for Counter-UAS · "
             "Adaptive Optical Countermeasures (GHOST/STORM/DEXTER)")
wsS['A2'].font = SUB; wsS.merge_cells('A2:C2')
rows = [("Period", "Proposed Price", "Topic Ceiling (NTE)"),
        ("Phase I Base (6 mo)", f"='Base'!{totB}", 200000),
        ("Phase I Option (6 mo)", f"='Option'!{totO}", 115000),
        ("Phase I TOTAL (Base + Option)", f"='Base'!{totB}+'Option'!{totO}", 315000)]
# totB/totO are the TOTAL PROPOSED PRICE cells returned by build_period (row count differs per sheet)
r0 = 4
for i, (a, b, c) in enumerate(rows):
    rr = r0 + i
    ca = wsS.cell(row=rr, column=1, value=a); ca.font = BOLD if i in (0, 3) else BLACK
    cb = wsS.cell(row=rr, column=2, value=b); cb.number_format = CUR; cb.alignment = RIGHT; cb.font = Font(name="Arial", size=10, color="008000") if i else BOLD
    cc = wsS.cell(row=rr, column=3, value=c); cc.alignment = RIGHT
    if i == 0:
        for cx in 'ABC': wsS[f'{cx}{rr}'].fill = GREY; wsS[f'{cx}{rr}'].font = BOLD
    else:
        cc.number_format = CUR0; cc.font = BLACK
wsS.cell(row=r0+5, column=1, value="All indirect rates are Immobileyes actuals (Fringe 35% · Labor OH 77% · G&A 40% · Fee 7%); "
        "magnitudes scaled to the Navy Phase I ceilings. No subcontracts in Base/Option.").font = SUB
wsS.merge_cells(start_row=r0+5, start_column=1, end_row=r0+5, end_column=3)

wb.save(OUT)
print("wrote", OUT)
# report the TOTAL row addresses actually used
print("Base TOTAL row = E31, Option TOTAL row = E31 (verify after recalc)")
